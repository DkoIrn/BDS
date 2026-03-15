"""Annotated dataset export (CSV and Excel) with QC flag columns."""

import io
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Severity priority for "worst severity" calculation
SEVERITY_PRIORITY = {"critical": 0, "warning": 1, "info": 2}

# Cell fills for Excel highlighting
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")


def _build_flag_map(issues: list[dict]) -> dict[int, dict]:
    """Build a mapping from 0-based row index to QC annotation data.

    Issues use 1-based row_number (matching the data rows, not header).
    We convert to 0-based index for DataFrame alignment.
    """
    row_map: dict[int, dict] = defaultdict(
        lambda: {"flag": "", "severity": "", "messages": []}
    )

    for issue in issues:
        row_num = issue.get("row_number")
        if row_num is None:
            continue
        idx = row_num - 1  # Convert 1-based to 0-based

        entry = row_map[idx]
        entry["flag"] = "FLAGGED"
        entry["messages"].append(str(issue.get("message", "")))

        # Track worst severity
        new_sev = str(issue.get("severity", "info"))
        current_sev = entry["severity"]
        if not current_sev or SEVERITY_PRIORITY.get(new_sev, 99) < SEVERITY_PRIORITY.get(current_sev, 99):
            entry["severity"] = new_sev

    return dict(row_map)


def export_annotated_csv(df: pd.DataFrame, issues: list[dict]) -> io.BytesIO:
    """Export DataFrame as CSV with QC annotation columns.

    Args:
        df: Original survey data.
        issues: List of validation issue dicts (row_number is 1-based).

    Returns:
        BytesIO containing CSV content.
    """
    annotated = df.copy()
    flag_map = _build_flag_map(issues)

    qc_flags = []
    qc_severities = []
    qc_messages = []

    for i in range(len(annotated)):
        entry = flag_map.get(i)
        if entry:
            qc_flags.append(entry["flag"])
            qc_severities.append(entry["severity"])
            qc_messages.append(" | ".join(entry["messages"]))
        else:
            qc_flags.append("")
            qc_severities.append("")
            qc_messages.append("")

    annotated["_qc_flag"] = qc_flags
    annotated["_qc_severity"] = qc_severities
    annotated["_qc_messages"] = qc_messages

    buf = io.BytesIO()
    annotated.to_csv(buf, index=False)
    buf.seek(0)
    return buf


def export_annotated_excel(df: pd.DataFrame, issues: list[dict]) -> io.BytesIO:
    """Export DataFrame as Excel with QC annotation columns and color-coded rows.

    Args:
        df: Original survey data.
        issues: List of validation issue dicts (row_number is 1-based).

    Returns:
        BytesIO containing Excel (.xlsx) content.
    """
    flag_map = _build_flag_map(issues)

    # Build annotated DataFrame
    annotated = df.copy()
    qc_flags = []
    qc_severities = []
    qc_messages = []

    for i in range(len(annotated)):
        entry = flag_map.get(i)
        if entry:
            qc_flags.append(entry["flag"])
            qc_severities.append(entry["severity"])
            qc_messages.append(" | ".join(entry["messages"]))
        else:
            qc_flags.append("")
            qc_severities.append("")
            qc_messages.append("")

    annotated["_qc_flag"] = qc_flags
    annotated["_qc_severity"] = qc_severities
    annotated["_qc_messages"] = qc_messages

    # Create workbook with openpyxl for styling control
    wb = Workbook()
    ws = wb.active
    ws.title = "QC Results"

    # Write header row
    headers = list(annotated.columns)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data rows with conditional formatting
    for row_idx, (_, row) in enumerate(annotated.iterrows(), start=2):
        data_index = row_idx - 2  # 0-based DataFrame index
        entry = flag_map.get(data_index)

        # Determine fill for this row
        fill = None
        if entry:
            severity = entry["severity"]
            if severity == "critical":
                fill = RED_FILL
            elif severity == "warning":
                fill = YELLOW_FILL

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if fill:
                cell.fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
