"""CSV and Excel parsers for the dispatch_parser registry."""

import csv
import io

from app.parsers.base import ParseResult


def parse_csv_file(file_bytes: bytes) -> ParseResult:
    """Parse CSV file bytes into a ParseResult.

    Tries UTF-8 with BOM first, then falls back to latin-1.
    """
    # Try utf-8-sig (handles BOM), fallback to latin-1
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1")

    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)

    if not all_rows:
        return ParseResult(
            headers=[],
            rows=[],
            total_rows=0,
            metadata={},
            source_format="csv",
        )

    headers = all_rows[0]
    rows = all_rows[1:]

    return ParseResult(
        headers=headers,
        rows=rows,
        total_rows=len(rows),
        metadata={"column_count": len(headers)},
        source_format="csv",
    )


def parse_excel_file(file_bytes: bytes) -> ParseResult:
    """Parse Excel (.xlsx/.xls) file bytes into a ParseResult.

    Uses openpyxl for reading. Reads the active sheet only.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    all_rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([str(cell) if cell is not None else "" for cell in row])

    wb.close()

    if not all_rows:
        return ParseResult(
            headers=[],
            rows=[],
            total_rows=0,
            metadata={},
            source_format="excel",
        )

    headers = all_rows[0]
    rows = all_rows[1:]

    return ParseResult(
        headers=headers,
        rows=rows,
        total_rows=len(rows),
        metadata={"column_count": len(headers)},
        source_format="excel",
    )
