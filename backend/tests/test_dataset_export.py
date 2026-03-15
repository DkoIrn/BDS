"""Tests for annotated dataset export (CSV and Excel)."""

import io

import openpyxl
import pandas as pd
import pytest

from app.services.dataset_export import export_annotated_csv, export_annotated_excel


@pytest.fixture
def source_df():
    """Simple DataFrame representing original survey data."""
    return pd.DataFrame(
        {
            "kp": [0.0, 0.01, 0.02, 0.03, 0.04],
            "dob": [1.2, 1.3, 15.0, 1.1, 1.4],
            "doc": [0.8, 0.9, 0.7, 0.8, 0.9],
        }
    )


@pytest.fixture
def issues_list():
    """Issues targeting rows 3 (critical) and 5 (warning), 1-based."""
    return [
        {
            "row_number": 3,
            "column_name": "dob",
            "severity": "critical",
            "message": "DOB value 15.0 exceeds maximum 10.0",
        },
        {
            "row_number": 5,
            "column_name": "doc",
            "severity": "warning",
            "message": "DOC value is a statistical outlier",
        },
    ]


class TestCSVExport:
    """Tests for export_annotated_csv."""

    def test_csv_adds_qc_columns(self, source_df, issues_list):
        result = export_annotated_csv(source_df, issues_list)
        df = pd.read_csv(result)
        assert "_qc_flag" in df.columns
        assert "_qc_severity" in df.columns
        assert "_qc_messages" in df.columns

    def test_csv_flags_correct_rows(self, source_df, issues_list):
        result = export_annotated_csv(source_df, issues_list)
        df = pd.read_csv(result)
        # Row 3 is 1-based -> index 2 in 0-based
        assert df.loc[2, "_qc_flag"] == "FLAGGED"
        assert df.loc[2, "_qc_severity"] == "critical"

    def test_csv_unflagged_rows_empty(self, source_df, issues_list):
        result = export_annotated_csv(source_df, issues_list)
        df = pd.read_csv(result)
        # Row 1 (index 0) has no issues
        assert pd.isna(df.loc[0, "_qc_flag"]) or df.loc[0, "_qc_flag"] == ""

    def test_csv_warning_severity(self, source_df, issues_list):
        result = export_annotated_csv(source_df, issues_list)
        df = pd.read_csv(result)
        # Row 5 is 1-based -> index 4 in 0-based
        assert df.loc[4, "_qc_flag"] == "FLAGGED"
        assert df.loc[4, "_qc_severity"] == "warning"


class TestExcelExport:
    """Tests for export_annotated_excel."""

    def test_excel_has_qc_columns(self, source_df, issues_list):
        result = export_annotated_excel(source_df, issues_list)
        wb = openpyxl.load_workbook(result)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "_qc_flag" in headers
        assert "_qc_severity" in headers
        assert "_qc_messages" in headers

    def test_excel_critical_rows_red_fill(self, source_df, issues_list):
        result = export_annotated_excel(source_df, issues_list)
        wb = openpyxl.load_workbook(result)
        ws = wb.active
        # Row 3 (1-based issue) -> data row index 2 (0-based) -> Excel row 4 (header=1, data starts at 2)
        # Actually: header row 1, data rows start at row 2. Row_number 3 -> 0-based index 2 -> Excel row 4
        critical_row = 4  # Excel row (1-based: header=1, data row 0=2, row 1=3, row 2=4)
        cell = ws.cell(row=critical_row, column=1)
        assert cell.fill.start_color.rgb is not None
        fill_color = str(cell.fill.start_color.rgb)
        assert "FFCCCC" in fill_color.upper() or "FFCCCC" in fill_color

    def test_excel_warning_rows_yellow_fill(self, source_df, issues_list):
        result = export_annotated_excel(source_df, issues_list)
        wb = openpyxl.load_workbook(result)
        ws = wb.active
        # Row 5 (1-based) -> 0-based index 4 -> Excel row 6
        warning_row = 6
        cell = ws.cell(row=warning_row, column=1)
        fill_color = str(cell.fill.start_color.rgb)
        assert "FFFFCC" in fill_color.upper() or "FFFFCC" in fill_color
